import openpyxl
from bs4 import BeautifulSoup as bs
import re
import http.cookiejar as cookielib
import cgi
import mechanize
import getpass
import argparse
import openpyxl
from progress.bar import Bar
import sys


BASE_URL = "https://aulaglobal.uc3m.es"
# Set the browser for the web crawler
br = mechanize.Browser()
cookiejar = cookielib.LWPCookieJar()
br.set_cookiejar(cookiejar)
br.set_handle_equiv(True)
br.set_handle_gzip(True)
br.set_handle_redirect(True)
br.set_handle_referer(True)
br.set_handle_robots(False)
br.set_handle_refresh(mechanize._http.HTTPRefreshProcessor(), max_time=1)
br.addheaders = [('User-agent',
                  'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.70 Safari/537.36')]
br.open(BASE_URL)

# Ask for NIA and password
print("####################################################\n" +
      "# Genera el Excel con los Correos de todo el mundo #\n" +
      "####################################################\n")

user = input("Enter NIA: ")
passwd = getpass.getpass(prompt="Enter password: ")

# Submit login form
print("Login in...")
br.select_form(nr=0)
br.form['adAS_username'] = user
br.form['adAS_password'] = passwd
br.submit(id="submit_ok")

# Check if success
url = br.open(BASE_URL)
login = url.get('X-Frame-Options', None)
status, _ = cgi.parse_header(login)
if status.upper() == "DENY":
    print("Login failed. Check your NIA and password and try again")
    exit(1)

path = sys.argv[1]
if path is None:
    path = input("Enter path to excel file: ")

wb = openpyxl.load_workbook(filename=path)
ws = wb['Sheet2']
links = []
email_list = []
for row in ws.iter_rows():
    for cell in row:
        if cell.column_letter == "B":
            links.append(cell.value)

counter = 1
with Bar('Processing...', max=len(links)) as bar:
    for target_url in links:
        if target_url is None or target_url == "0":
            counter += 1
            pass
        url = br.open(target_url)
        soup = bs(url, "html.parser")
        email_pattern = r"[0-9]+@alumnos.uc3m.es"

        emails = re.findall(email_pattern, soup.text)
        if emails is not None:
            for email in emails:
                email_list.append(email)
        bar.next()
        wb_1 = openpyxl.load_workbook(filename=path)
        ws_1 = wb_1['Sheet1']
        ws_1.cell(row=counter, column=2).value = email
        counter += 1
        wb_1.save(path)

print(email_list)
with open('emails.txt', 'w') as f:
    for item in email_list:
        f.write("%s\n" % item)
